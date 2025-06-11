from playwright.sync_api import sync_playwright
import configparser
import pandas as pd
import time
from automation.login import LoginMicrosoft
import os

def get_settings():
    try:
        from django.conf import settings
        _ = settings.CONFIG_PATH
        return settings
    except Exception:
        class SimpleSettings:
            CONFIG_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../conf.cfg'))
            CREDENTIALS_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../credentials.cfg'))
            EXCEL_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../../TeamsDrive/Respostas.xlsx'))
        return SimpleSettings()

settings = get_settings()

def preencher_pergunta_texto(page, texto_pergunta, resposta):
    pergunta = page.locator(f"//span[contains(text(), '{texto_pergunta}')]")
    input_field = pergunta.locator("xpath=following::input[1] | following::textarea[1]").first

    input_field.wait_for(state="visible", timeout=10000)
    input_field.fill(resposta)

def preencher_pergunta_multipla_escolha(page, texto_pergunta, resposta):
    pergunta = page.locator(f"//span[contains(text(), '{texto_pergunta}')]")
    opcao = pergunta.locator(f"xpath=following::label//span[contains(text(), '{resposta}')]").first
    
    opcao.wait_for(state="visible", timeout=10000)
    opcao.click()

def preencher_pergunta_multipla_outro(page, texto_pergunta, opcoes, resposta):
    pergunta = page.get_by_label(texto_pergunta)
    pergunta.scroll_into_view_if_needed()
    pergunta.wait_for(state="visible", timeout=10000)

    if resposta in opcoes:
        pergunta.get_by_role("radio", name=resposta).click()
    else:
        pergunta.get_by_role("radio", name="Other answer").click()
        input_outro = pergunta.locator("input[aria-label='Other answer'][data-automation-id='textInput']")
        input_outro.wait_for(state="visible", timeout=10000)
        input_outro.fill(str(resposta))


def run():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, slow_mo=100) # headless=True quando usar na VM
                                                                # headless=False para demonstrar

        config = configparser.ConfigParser()
        config.read(settings.CONFIG_PATH)
        url = config['START']['forms_admin_url'] + "&topview=Prefill"
        excel_path = settings.EXCEL_PATH

        dados = pd.read_excel(excel_path, 'Respostas')
        dados = dados.loc[dados.groupby('Email Institucional')['Ano de submissão'].idxmax()]
        matriz = dados.values.tolist()

        emails_autorizados = pd.read_excel(excel_path, sheet_name='Emails_Autorizados')

        config.read(settings.CREDENTIALS_PATH)
        email = config['CREDENTIALS']['email']
        password = config['CREDENTIALS']['password']

        page = browser.new_page()
        page.goto(url)

        try:
            login_tool = LoginMicrosoft()
            login_tool.login(page, email, password)

            for linha in matriz:
                email_pessoal = linha[2] if pd.notna(linha[2]) else linha[0]
                email_institucional = linha[1]
                temEmprego = linha[3]
                empresa = linha[4]
                pais = linha[5]
                posicao = linha[6]
                lideranca = linha[7]
                areaLicenc = linha[8]
                gamaSalario = linha[9]
                faltaEmprego = linha[10]
                numMeses = linha[11]
                
                page.goto(url)
                # Preencher campos
                page.locator('input').nth(0).wait_for(state="visible")
                page.fill('input >> nth=0', email_institucional[0:7])
                page.fill('input >> nth=1', email_pessoal)
                
                page.get_by_text("Não", exact=True).nth(0).click()

                page.get_by_text(temEmprego, exact=True).nth(1).click()

                time.sleep(3)

                if(temEmprego=="Sim"):
                    preencher_pergunta_texto(page, "Que empresa que o contratou?", empresa)
                    preencher_pergunta_texto(page, "Em que país é que foi contratado?", pais)
                    preencher_pergunta_texto(page, "Que posição é que ocupa nesse emprego?", posicao)
                    preencher_pergunta_multipla_escolha(page, "É uma posição de liderança", lideranca)
                    preencher_pergunta_multipla_escolha(page, "Foi na área em que se licenciou?", areaLicenc)
                    preencher_pergunta_multipla_escolha(page, "Em qual destas gamas é que o seu salário se encontra?", gamaSalario)
                else:
                    preencher_pergunta_multipla_outro(page, "Porque é que não tem emprego?", ["Não procurei", "Procurei, mas não fui aceite"], faltaEmprego)

                preencher_pergunta_multipla_outro(page, "meses que demorou", ["Ainda não tive"], numMeses)

                
                page.get_by_role("button", name="Get Prefilled Link" # ou "Obter Ligação Pré-Preenchida" em PT
                                ).click()


                page.get_by_role("button", name="Copy link" # ou "Copiar ligação" em PT
                                ).wait_for(state="visible")
                page.get_by_role("button", name="Copy link").click()

                link_input = page.locator('input[readonly][value^="http"]').first
                link_input.wait_for(state="visible")
                link_copiado = link_input.input_value()

                print("Link copiado:", link_copiado)
                mask = emails_autorizados['Email Institucional'] == email
                emails_autorizados.loc[mask, 'Link a enviar'] = link_copiado

            from openpyxl import load_workbook
            from openpyxl.utils.dataframe import dataframe_to_rows

            wb = load_workbook(excel_path)
            if 'Emails_Autorizados' in wb.sheetnames:
                ws = wb['Emails_Autorizados']
                ws.delete_rows(2, ws.max_row)  # Preserva a primeira linha (cabeçalhos)

                for r_idx, row in enumerate(dataframe_to_rows(emails_autorizados, index=False, header=False), start=2):
                    for c_idx, value in enumerate(row, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
            else:
                ws = wb.create_sheet('Emails_Autorizados')
                for row in dataframe_to_rows(emails_autorizados, index=False, header=True):
                    ws.append(row)

            wb.save(excel_path)
        
            browser.close()
            return "Operação terminada com sucesso"

        except Exception as e:
            page.screenshot(path="Error.png")
            browser.close()
            return f"Erro ao executar operação: {str(e)}"

if __name__ == "__main__":
    run()
